"""
工具函数
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np


def generate_years(construction_period, operation_period):
    """
    根据建设期和运营期生成年份列表

    Args:
        construction_period: 建设期（年）
        operation_period: 运营期（年）

    Returns:
        list: 年份列表，如 ["第1年", "第2年", ..., "第20年"]
    """
    total_period = construction_period + operation_period
    return [f"第{i+1}年" for i in range(total_period)]


def is_yellow_cell(cell):
    """
    判断单元格是否为黄色背景

    Args:
        cell: openpyxl单元格对象

    Returns:
        bool: 是否为黄色背景
    """
    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
        # 黄色RGB值可能为 "FFFFFF00" 或 "FFFF00"
        bg_color = cell.fill.start_color.rgb
        return bg_color in ["FFFFFF00", "FFFF00", "0000FFFF"]  # 不同版本openpyxl的颜色表示
    return False


def load_excel_with_formulas(file_path, sheet_name=None):
    """
    加载Excel文件，保留公式

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，如果为None则返回所有工作表

    Returns:
        如果指定sheet_name，返回DataFrame；否则返回字典
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)

    if sheet_name:
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                # 优先使用公式，如果没有公式则使用值
                row_data.append(cell.value)
            data.append(row_data)
        wb.close()
        return pd.DataFrame(data)
    else:
        result = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = []
            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    row_data.append(cell.value)
                data.append(row_data)
            result[sheet_name] = pd.DataFrame(data)
        wb.close()
        return result


def load_excel_with_colors(file_path, sheet_name):
    """
    加载Excel文件，并标记黄色单元格

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称

    Returns:
        tuple: (DataFrame, 标记黄色的DataFrame)
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # 数据
    data = []
    yellow_mask = []

    for row in ws.iter_rows():
        row_data = []
        row_yellow = []
        for cell in row:
            row_data.append(cell.value)
            row_yellow.append(is_yellow_cell(cell))
        data.append(row_data)
        yellow_mask.append(row_yellow)

    wb.close()

    df = pd.DataFrame(data)
    yellow_df = pd.DataFrame(yellow_mask)

    return df, yellow_df


def compare_with_original(calculated_data, original_data, tolerance=0.01):
    """
    对比计算结果与原始数据

    Args:
        calculated_data: 计算得到的DataFrame
        original_data: 原始DataFrame
        tolerance: 容差阈值（相对误差）

    Returns:
        DataFrame: 包含差异信息的DataFrame
    """
    result = pd.DataFrame(index=calculated_data.index)

    for col in calculated_data.columns:
        if col in original_data.columns:
            calc_vals = pd.to_numeric(calculated_data[col], errors='coerce')
            orig_vals = pd.to_numeric(original_data[col], errors='coerce')

            diff = calc_vals - orig_vals
            # 避免除以0
            rel_diff = np.where(orig_vals.abs() > 0.01, diff / orig_vals.abs(), diff)

            result[f"{col}_diff"] = diff
            result[f"{col}_rel_diff"] = rel_diff

            # 标记差异超过容差的单元格
            result[f"{col}_flag"] = rel_diff.abs() > tolerance

    return result


def format_number(value, decimal_places=2):
    """
    格式化数字

    Args:
        value: 数值
        decimal_places: 小数位数

    Returns:
        格式化后的字符串
    """
    if pd.isna(value) or value is None:
        return ""
    if isinstance(value, (int, float)):
        return f"{value:,.{decimal_places}f}"
    return str(value)


def get_input_module_ranges():
    """
    定义输入模块的行范围

    Returns:
        dict: 模块名称到行范围的映射
    """
    return {
        "1. 基础信息": (1, 2),
        "2. 项目投资": (9, 28),
        "3. 资产形成": (32, 45),
        "4. 资产销售计划": (47, 54),
        "5. 投融资计划": (59, 93),
        "6. 银行借款计划": (98, 113),
        "7. 产品销售收入": (114, 145),
        "8. 外购材料成本": (146, 202),
        "9. 外购燃料及动力": (203, 210),
        "10. 工资福利成本": (211, 230),
        "11. 修理费及其他费用": (231, 245),
        "12. 销售费用": (246, 260),
    }


def safe_divide(numerator, denominator):
    """
    安全除法，避免除以0

    Args:
        numerator: 分子
        denominator: 分母

    Returns:
        除法结果，如果分母为0则返回0
    """
    if pd.isna(denominator) or denominator == 0:
        return 0
    return numerator / denominator


def calculate_pv(cash_flow, discount_rate, year):
    """
    计算现值

    Args:
        cash_flow: 现金流
        discount_rate: 折现率
        year: 年份

    Returns:
        现值
    """
    return cash_flow / ((1 + discount_rate) ** year)
